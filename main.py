import argparse
import json
import os
import typing

import gspread
import openpyxl
from openpyxl.cell import Cell
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from src.games import games
from src.base import players, Player, Run

OUTPUT_FILE = "codpointsoutput/output.xlsx"


def _autosize_column(ws: Worksheet):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)) + 5)
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def _dump_totals_sheet(wb: openpyxl.Workbook):
    _players = sorted(
        players.values(), key=lambda player: player.total_points, reverse=True
    )
    ws = typing.cast(Worksheet, wb.active)
    ws.title = "Totals"
    ws.append(["Rank", "Player", "Points"])

    tl_border = Border(
        left=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
    )
    t_border = Border(
        top=Side(border_style="medium", color="000000"),
    )
    tr_border = Border(
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
    )
    left_border = Border(
        left=Side(border_style="medium", color="000000"),
    )
    right_border = Border(
        right=Side(border_style="medium", color="000000"),
    )
    bl_border = Border(
        left=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000"),
    )
    b_border = Border(
        bottom=Side(border_style="medium", color="000000"),
    )
    br_border = Border(
        right=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000"),
    )

    i = 0
    _prev_A = None
    _prev_B = None
    _prev_C = None
    specialist_start = False
    veteran_start = False
    hardened_start = False
    regular_start = False
    recruit_start = False
    end_start = False

    for i, player in enumerate(_players):
        A = Cell(ws, value=f"{i + 1}")  # type: ignore
        B = Cell(ws, value=f"{player.name}")  # type: ignore
        C = Cell(ws, value=f"{player.total_points}")  # type: ignore

        ws.append([A, B, C])

        if player.total_points >= 80000:  # Specialist
            if not specialist_start:
                A.border = tl_border
                B.border = t_border
                C.border = tr_border
            else:
                A.border = left_border
                C.border = right_border

            for cell in ws[i + 2]:
                cell.fill = PatternFill(fgColor="2ecc70", fill_type="solid")

            specialist_start = True
        elif player.total_points >= 40000:  # Veteran
            if not veteran_start:
                A.border = tl_border
                B.border = t_border
                C.border = tr_border
                # Retroactively put the bottom border on the previous row
                assert (
                    _prev_A is not None and _prev_B is not None and _prev_C is not None
                ), f"{i}, {player}"
                _prev_A.border = bl_border
                _prev_B.border = b_border
                _prev_C.border = br_border
            else:
                A.border = left_border
                C.border = right_border

            for cell in ws[i + 2]:
                cell.fill = PatternFill(fgColor="ff00db", fill_type="solid")

            veteran_start = True
        elif player.total_points >= 20000:  # Hardened
            if not hardened_start:
                A.border = tl_border
                B.border = t_border
                C.border = tr_border
                # Retroactively put the bottom border on the previous row
                assert (
                    _prev_A is not None and _prev_B is not None and _prev_C is not None
                )
                _prev_A.border = bl_border
                _prev_B.border = b_border
                _prev_C.border = br_border
            else:
                A.border = left_border
                C.border = right_border

            for cell in ws[i + 2]:
                cell.fill = PatternFill(fgColor="3498db", fill_type="solid")
            hardened_start = True
        elif player.total_points >= 10000:  # Regular
            # Then continue with this one
            if not regular_start:
                A.border = tl_border
                B.border = t_border
                C.border = tr_border
                # Retroactively put the bottom border on the previous row
                assert (
                    _prev_A is not None and _prev_B is not None and _prev_C is not None
                )
                _prev_A.border = bl_border
                _prev_B.border = b_border
                _prev_C.border = br_border
            else:
                A.border = left_border
                C.border = right_border

            for cell in ws[i + 2]:
                cell.fill = PatternFill(fgColor="e67e22", fill_type="solid")
            regular_start = True
        elif player.total_points >= 5000:  # Recruit
            # Then continue with this one
            if not recruit_start:
                A.border = tl_border
                B.border = t_border
                C.border = tr_border
                # Retroactively put the bottom border on the previous row
                assert (
                    _prev_A is not None and _prev_B is not None and _prev_C is not None
                )
                _prev_A.border = bl_border
                _prev_B.border = b_border
                _prev_C.border = br_border
            else:
                A.border = left_border
                C.border = right_border

            for cell in ws[i + 2]:
                cell.fill = PatternFill(fgColor="9b59b6", fill_type="solid")
            recruit_start = True
        else:
            if not end_start:
                A.border = tl_border
                B.border = t_border
                C.border = tr_border
                # Retroactively put the bottom border on the previous row
                assert (
                    _prev_A is not None and _prev_B is not None and _prev_C is not None
                )
                _prev_A.border = bl_border
                _prev_B.border = b_border
                _prev_C.border = br_border
            else:
                A.border = left_border
                C.border = right_border

            end_start = True

        _prev_A = A
        _prev_B = B
        _prev_C = C

    # At the end, put the bottom border on the last row
    for cell in ws[i + 2]:
        cell.border = Border(
            left=Side(border_style="medium", color="000000"),
            right=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000"),
        )

    ws.freeze_panes = "A2"
    _autosize_column(ws)


def _dump_main_categories_sheet(wb: openpyxl.Workbook):
    _players = sorted(
        players.values(), key=lambda player: player.main_points, reverse=True
    )

    ws: Worksheet = wb.create_sheet("Main Categories")
    ws.append(["Rank", "Player", "Points"])
    for i, player in enumerate(_players):
        ws.append([i + 1, player.name, player.main_points])
    ws.freeze_panes = "A2"
    _autosize_column(ws)


def _dump_il_categories_sheet(wb: openpyxl.Workbook):
    _players = sorted(
        players.values(), key=lambda player: player.il_points, reverse=True
    )

    ws: Worksheet = wb.create_sheet("IL")
    ws.append(["Rank", "Player", "Points"])
    for i, player in enumerate(_players):
        ws.append([i + 1, player.name, player.il_points])
    ws.freeze_panes = "A2"
    _autosize_column(ws)


def _dump_game_breakdown_sheet(wb: openpyxl.Workbook):
    _players = sorted(
        players.values(), key=lambda player: player.total_points, reverse=True
    )

    ws: Worksheet = wb.create_sheet("Game Breakdown")
    headers = ["Rank", "Player", "Points"]
    for game in games:
        headers.append(game.game.upper())
    ws.append(headers)

    for i, player in enumerate(_players):
        row = [i + 1, player.name, player.total_points]
        for game in games:
            row.append(sum(run.points for run in player.runs if run.game == game.game))
        ws.append(row)

    ws.freeze_panes = "A2"
    _autosize_column(ws)


def _dump_runs_sheet(wb: openpyxl.Workbook):
    _players = sorted(
        players.values(), key=lambda player: player.total_points, reverse=True
    )

    ws: Worksheet = wb.create_sheet("Runs")
    ws.append(["Player", "Game", "Run", "Place", "Time", "Points"])
    for player in _players:
        for run in player.runs:
            ws.append(
                [player.name, run.game, run.name, run.place, run.time, run.points]
            )
    ws.freeze_panes = "A2"
    _autosize_column(ws)


def _create_charts(wb: openpyxl.Workbook):
    _game_data: typing.Dict[str, int] = {}
    game_order = [game.game for game in games]

    for player in players.values():
        for run in player.runs:
            _run_count = _game_data.get(run.game, 0)
            _game_data[run.game] = _run_count + 1

    # Sort the game data by the order of the games
    sorted_game_data = sorted(
        _game_data.items(), key=lambda item: game_order.index(item[0])
    )

    # Create raw sheet for referencing data
    raw_sheet: Worksheet = wb.create_sheet("Raw")
    raw_sheet.append(["Game", "Runs"])
    for game, runs in sorted_game_data:
        raw_sheet.append([game, runs])

    chart = BarChart()
    chart.title = "Runs by Game"
    labels = Reference(raw_sheet, min_col=1, min_row=2, max_row=len(_game_data) + 1)
    data = Reference(raw_sheet, min_col=2, min_row=1, max_row=len(_game_data) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    wb["Totals"].add_chart(chart, "G9")  # type: ignore


def from_json():
    with open("runs.json", "r") as f:
        runs = json.load(f)

    for player in runs["players"]:
        players[player["id"]] = Player(player["id"], player["name"], [])

        for run in player["runs"]:
            players[player["id"]].runs.append(Run(**run))


def to_json():
    # In case I want to do something with the runs later, serialize them
    data = {
        "players": [
            {
                "id": player.id,
                "name": player.name,
                "runs": [run.__dict__ for run in player.runs],
            }
            for player in players.values()
        ]
    }
    with open("runs.json", "w") as f:
        json.dump(data, f, indent=4)


def dump():
    wb = openpyxl.Workbook()
    _dump_totals_sheet(wb)
    _dump_main_categories_sheet(wb)
    _dump_il_categories_sheet(wb)
    _dump_runs_sheet(wb)
    _dump_game_breakdown_sheet(wb)
    _create_charts(wb)

    wb.save(OUTPUT_FILE)


def upload():
    if not os.path.exists("token.json"):
        raise Exception("token.json not found")

    gc = gspread.service_account(filename="token.json")  # type: ignore
    sh = gc.open_by_key("1GDZOL2nklQGk6D4hND43sd3s3-4qLHM3GFHeeX54-ZM")

    file = pd.ExcelFile(OUTPUT_FILE)

    for name in file.sheet_names:
        assert isinstance(name, str)
        data = file.parse(name)

        worksheet = sh.worksheet(name)
        worksheet.clear()
        worksheet.update([data.columns.values.tolist()] + data.values.tolist())


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--json",
        action="store_true",
        help="Use the runs.json file instead of the spreadsheet",
    )

    args = parser.parse_args()
    if args.json:
        from_json()
    else:
        for game in games:
            game().calculate()

        to_json()

    dump()
    upload()

    print("Done!")
